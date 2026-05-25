"""
Transcript Parser — Benha University
Returns all parsed data as a JSON-serializable dict.
Usage: python parser.py transcript.xlsx
"""

import json
import math
import re
import sys
import pandas as pd
from openpyxl import load_workbook

# ── Column positions (1-indexed) ──────────────────────────────────────────────
COL_PTS_X_HRS = 9
COL_GRADE_PTS = 16
COL_GRADE     = 22
COL_SCORE     = 28
COL_CREDITS   = 32
COL_COURSE    = 40
COL_HEADER    = 9
COL_SEASON    = 16
COL_STATS_CUM = 16
COL_STATS_SEM = 40

# ── Student info cell positions (row, col) ────────────────────────────────────
STUDENT_FIELDS = {
    'name':          (4,  18),
    'id':            (6,  19),
    'level':         (11, 11),
    'faculty':       (13,  4),
    'university':    (14, 50),
    'overall_grade': (20, 18),
    'passed_hours':  (25, 23),
    'national_id':   (32, 39),
    'dob':           (34, 45),
    'nationality':   (38, 43),
}

GRADE_PASS = {'A+','A','A-','B+','B','B-','C+','C','C-','D+','D'}
SEM_ORDER  = {'S1': 0, 'S2': 1, 'Summer': 2}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _cell(ws, row, col):
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None

def _row_cell(row_cells, col):
    idx = col - 1
    if idx < 0 or idx >= len(row_cells):
        return None
    return row_cells[idx].value

def _str(v):
    return str(v).strip() if v is not None else None

def _float(v):
    try:    return float(v) if v is not None else None
    except: return None

def _int(v):
    try:    return int(float(v)) if v is not None else None
    except: return None

def sanitize(obj):
    """Recursively replace float NaN/Inf with None for JSON safety."""
    if isinstance(obj, list):
        return [sanitize(i) for i in obj]
    if isinstance(obj, dict):
        return {k: sanitize(v) for k, v in obj.items()}
    try:
        if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
    except Exception:
        pass
    return obj

# ── Row classifiers ───────────────────────────────────────────────────────────

def is_level_header(row_cells):
    v = _row_cell(row_cells, COL_HEADER)
    if not v:
        return False
    s = str(v)
    return ('المستوى' in s or 'المستوي' in s) and re.search(r'\d{4}-\d{4}', s)

def is_season_header(row_cells):
    v = _row_cell(row_cells, COL_SEASON)
    if not v:
        return False
    return 'المستوى / الفصل' in str(v)

def is_stats_row(row_cells):
    v = _row_cell(row_cells, COL_STATS_CUM)
    if not v:
        return False
    return 'التراكميه' in str(v)

def is_course_row(row_cells):
    pts  = _row_cell(row_cells, COL_PTS_X_HRS)
    name = _row_cell(row_cells, COL_COURSE)
    if pts is None or name is None:
        return False
    try:
        float(pts)
    except (TypeError, ValueError):
        return False
    name_str = str(name).strip()
    return name_str not in ('', 'اسم المقرر', 'nan')

# ── Parsers ───────────────────────────────────────────────────────────────────

def parse_level_header(text):
    level_map = {
        'الأول':   1, 'الاول':   1,
        'الثاني':  2, 'الثانى':  2,
        'الثالث':  3, 'الثالثة': 3,
        'الرابع':  4, 'الرابعة': 4,
        'الأولى':  1,
    }
    level = next((num for ar, num in level_map.items() if ar in text), None)
    m     = re.search(r'(\d{4}-\d{4})', text)
    year  = m.group(1) if m else None
    m2    = re.search(r'تراكمى نقاط المستوى[:\-\s]+([\d.]+)', text)
    cgpa  = float(m2.group(1)) if m2 else None
    return level, year, cgpa

def parse_season_header(text):
    if 'الصيف'  in text:
        return 'Summer'
    if 'الثاني' in text or 'الثانى' in text:
        return 'S2'
    return 'S1'

def parse_stats_line(cum_text, sem_text):
    cgpa = sgpa = None
    if cum_text:
        m = re.search(r'النقاط التراكميه[:\-\s]+([\d.]+)', str(cum_text))
        if m:
            cgpa = float(m.group(1))
    if sem_text:
        m = re.search(r'النقاط الفصلية[:\-\s]+([\d.]+)', str(sem_text))
        if m:
            sgpa = float(m.group(1))
    return cgpa, sgpa

def determine_status(grade_str, score):
    if grade_str == 'غياب':         return 'Absent'
    if grade_str == 'راسب لائحة':   return 'Fail (Regulation)'
    if grade_str == 'F':            return 'Fail'
    if grade_str in GRADE_PASS:     return 'Pass'
    if score is not None:           return 'Pass' if score >= 50 else 'Fail'
    return 'Fail'

# ── Student info ──────────────────────────────────────────────────────────────

def parse_student_info(ws):
    def get(row, col):
        return _cell(ws, row, col)

    raw_grade = get(*STUDENT_FIELDS['overall_grade'])
    if raw_grade:
        m = re.search(r'التقدير العام\s+(.+)', str(raw_grade))
        overall_grade = m.group(1).strip() if m else str(raw_grade).strip()
    else:
        overall_grade = None

    # Cumulative GPA = last occurrence of النقاط التراكميه in the sheet
    last_cgpa = None
    for sheet_row in ws.iter_rows(values_only=False):
        for cell in sheet_row:
            if cell.value and isinstance(cell.value, str) and 'النقاط التراكميه' in cell.value:
                m = re.search(r'النقاط التراكميه[:\-\s]+([\d.]+)', cell.value)
                if m:
                    last_cgpa = float(m.group(1))

    raw_dob = get(*STUDENT_FIELDS['dob'])
    try:
        dob = pd.Timestamp(raw_dob).strftime('%Y-%m-%d') if raw_dob else None
    except Exception:
        dob = str(raw_dob) if raw_dob else None

    raw_hours = get(*STUDENT_FIELDS['passed_hours'])
    try:
        passed_hours = int(float(raw_hours)) if raw_hours else None
    except Exception:
        passed_hours = None

    return {
        "name":           _str(get(*STUDENT_FIELDS['name'])),
        "student_id":     _str(get(*STUDENT_FIELDS['id'])),
        "current_level":  _str(get(*STUDENT_FIELDS['level'])),
        "faculty":        _str(get(*STUDENT_FIELDS['faculty'])),
        "university":     _str(get(*STUDENT_FIELDS['university'])),
        "overall_grade":  overall_grade,
        "cumulative_gpa": last_cgpa,
        "passed_hours":   passed_hours,
        "national_id":    _str(get(*STUDENT_FIELDS['national_id'])),
        "date_of_birth":  dob,
        "nationality":    _str(get(*STUDENT_FIELDS['nationality'])),
    }

# ── Main parser ───────────────────────────────────────────────────────────────

def parse_transcript(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    student_info = parse_student_info(ws)

    courses = []
    cur_level = cur_year = cur_season = cur_cgpa = cur_sgpa = None

    for sheet_row in ws.iter_rows(values_only=False):
        row_cells = list(sheet_row)

        if is_level_header(row_cells):
            text = str(_row_cell(row_cells, COL_HEADER))
            cur_level, cur_year, gpa = parse_level_header(text)
            if gpa:
                cur_cgpa = gpa
            cur_season = None
            continue

        if is_season_header(row_cells):
            cur_season = parse_season_header(str(_row_cell(row_cells, COL_SEASON)))
            continue

        if is_stats_row(row_cells):
            cgpa, sgpa = parse_stats_line(
                _row_cell(row_cells, COL_STATS_CUM),
                _row_cell(row_cells, COL_STATS_SEM)
            )
            if cgpa:
                cur_cgpa = cgpa
            if sgpa:
                cur_sgpa = sgpa
            continue

        if is_course_row(row_cells):
            grade_str = _str(_row_cell(row_cells, COL_GRADE)) or ''
            score_raw = _row_cell(row_cells, COL_SCORE)
            score = None
            if score_raw is not None and str(score_raw).strip() not in ('*', ''):
                try:
                    score = float(score_raw)
                except Exception:
                    pass

            courses.append({
                "level":          cur_level,
                "academic_year":  cur_year,
                "semester":       cur_season,
                "course_name":    _str(_row_cell(row_cells, COL_COURSE)),
                "credits":        _int(_row_cell(row_cells, COL_CREDITS)),
                "score":          score,
                "grade":          grade_str,
                "grade_points":   _float(_row_cell(row_cells, COL_GRADE_PTS)),
                "pts_x_hrs":      _float(_row_cell(row_cells, COL_PTS_X_HRS)),
                "status":         determine_status(grade_str, score),
                "cumulative_gpa": cur_cgpa,
                "semester_gpa":   cur_sgpa,
            })

    # Deduplicate and sort
    df = pd.DataFrame(courses)
    if not df.empty:
        df = df.drop_duplicates(
            subset=['level', 'academic_year', 'semester', 'course_name', 'score']
        )
        df['_ord'] = df['semester'].map(SEM_ORDER).fillna(9)
        df = (
            df.sort_values(['level', 'academic_year', '_ord'])
            .drop(columns='_ord')
            .reset_index(drop=True)
        )
        courses = df.to_dict(orient='records')

    return sanitize({
        "student_info": student_info,
        "courses":      courses,
    })


# ── CLI entry point ───────────────────────────────────────────────────────────

if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else '/mnt/user-data/uploads/Book1.xlsx'
    result = parse_transcript(path)
    print(json.dumps(result, indent=2, ensure_ascii=False))